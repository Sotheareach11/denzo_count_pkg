"""
PKG Counter + Combine PDFs - Web App (Flask)
=============================================
Runs in the browser - works great on CodeSandbox, Replit, Render, or any
container with no display (no tkinter needed).

Install:
    pip install flask pdfplumber openpyxl pypdf gunicorn

Run locally:
    python app.py

Run in production (Render Start Command):
    gunicorn app:app --workers 1 --threads 4 --timeout 180 --bind 0.0.0.0:$PORT

    NOTE: On a low-RAM Render plan (512MB-1GB), use --workers 1. Each
    gunicorn worker loads a full separate copy of the app + its libraries
    (pdfplumber, openpyxl, etc.) into memory - 2 workers roughly doubles
    baseline memory use before you've even processed a single PDF. With
    threads instead of workers you still get concurrency without the
    memory multiplication.

Two tools in this app:
  1. PKG Counter ("/")        - upload packing list (PKL) PDFs and/or
                                 invoice (INV) PDFs together in one batch.
                                 Each file is auto-detected by type. PKL
                                 files are parsed exactly as before (CML
                                 No. / PKG / weight breakdown). When a PKL
                                 and an INV are for the SAME shipment
                                 (matched by the Invoice No. printed on
                                 each PDF, falling back to the numeric ID
                                 in the filename if a PKL doesn't have
                                 that field), they're merged into ONE
                                 sheet: the normal packing-list sheet, with
                                 a Unit Price / Amount column added per
                                 item from the matching invoice. If an INV
                                 has no matching PKL, it still gets its own
                                 sheet so nothing is lost. The "PKG
                                 Summary" sheet is likewise one row-set per
                                 shipment (not duplicated per file),
                                 grouped by Ship To (factory), with an
                                 Amount (USD) column carrying each
                                 shipment's total invoiced value.
  2. Combine PDFs ("/combine") - upload INV / PL / Freight PDFs, get them
                                 merged into one PDF, grouped by invoice
                                 number and ordered INV -> PL -> Freight.
"""

import os
import re
import io
import gc
import uuid
import shutil
import threading
from collections import OrderedDict
from combinepdf import combine

from flask import Flask, request, render_template, send_file, jsonify
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# PRODUCTION / RENDER SETTINGS
# ---------------------------------------------------------------------------

# Cap total request size (all uploaded files combined) so a huge batch of
# PDFs can't exhaust memory or blow past Render's proxy limits. Adjust the
# number (in MB) to whatever your plan's RAM comfortably allows.
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "75"))
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024


@app.errorhandler(413)
def too_large(e):
    """Return a clean JSON error instead of a raw connection failure when
    the upload exceeds MAX_CONTENT_LENGTH."""
    return jsonify({
        "error": f"Upload too large. Please upload no more than "
                 f"{MAX_UPLOAD_MB}MB of PDFs at a time (try fewer files "
                 f"or split into batches)."
    }), 413


# ---------------------------------------------------------------------------
# BACKGROUND JOB STORE (PKG Counter)
# ---------------------------------------------------------------------------
# Large batches of PDFs can take longer to process than Render's/gunicorn's
# request timeout allows. Rather than making the browser hold one huge HTTP
# request open the whole time (which is what was timing out), /count now
# returns almost immediately with a job_id, a background thread does the
# actual parsing, and the browser polls /count/status/<job_id> for progress.
#
# JOBS is a plain in-memory dict. This is fine as long as the app runs as a
# single gunicorn worker (see the Start Command note above) - if you ever
# scale to multiple workers/instances, move this to Redis or a database so
# all workers can see the same job state.
JOBS = {}
JOBS_LOCK = threading.Lock()


def _set_job(job_id, **kwargs):
    with JOBS_LOCK:
        JOBS[job_id].update(kwargs)


def _process_count_job(job_id, saved_files):
    """
    Runs in a background thread. saved_files is a list of
    (temp_path, original_filename) tuples already written to disk by the
    /count route (so the HTTP request itself can return immediately).

    Two-phase processing:
      1. Parse every file individually (PKL or INV, auto-detected) into a
         lightweight "parsed" dict - this never fails because two files
         are related to each other, only because a single file is
         unreadable.
      2. build_merged_entries() groups those parsed files by shipment
         (Invoice No., or a fallback ID from the filename) and merges any
         matching PKL + INV pair into one entry before Excel/preview
         building happens.
    """
    parsed_files = []
    preview_errors = {}
    total = len(saved_files)

    try:
        for idx, (temp_path, filename) in enumerate(saved_files, start=1):
            _set_job(job_id, status="processing",
                     progress_done=idx - 1, progress_total=total,
                     current_file=filename)
            try:
                doc_type = detect_doc_type(temp_path)
                if doc_type == "invoice":
                    records, header = parse_invoice(temp_path)
                    parsed_files.append({
                        "type": "invoice",
                        "filename": filename,
                        "ship_to": header["ship_to"],
                        "invoice_no": header["invoice_no"],
                        "records": records,
                    })
                else:
                    records, packages, ship_to, invoice_no = parse_packing_list(
                        temp_path)
                    summary, package_names = summarize_by_name(records)
                    parsed_files.append({
                        "type": "packing_list",
                        "filename": filename,
                        "ship_to": ship_to,
                        "invoice_no": invoice_no,
                        "summary": summary,
                        "packages": packages,
                        "package_names": package_names,
                    })
            except Exception as file_err:
                # Don't let one bad/unreadable PDF kill the whole batch -
                # record it and keep going.
                preview_errors[filename] = str(file_err)
            finally:
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
                # Release pdfplumber's internal buffers for this file
                # before moving on, so memory doesn't climb across a big
                # batch.
                gc.collect()

        _set_job(job_id, progress_done=total)

        all_results = build_merged_entries(parsed_files)

        if not all_results and not preview_errors:
            _set_job(job_id, status="error",
                     error="No valid PDF data extracted.")
            return

        preview = []
        for filename, err in preview_errors.items():
            preview.append({"filename": filename, "error": err})

        grand_pkg = 0
        grand_nw = 0.0
        grand_gw = 0.0
        grand_amount = 0.0

        for filename, summary, packages, package_names, ship_to, invoice_amount in all_results:
            total_nw = sum(p["nw"] for p in packages.values())
            total_gw = sum(p["gw"] for p in packages.values())

            preview.append({
                "filename": filename,
                "ship_to": ship_to,
                "items": [
                    {"itemno": r["itemno"],
                        "desc": display_name(r["custitem"], r["desc"]),
                        "pkg": r["count"], "qty": r["total_qty"],
                        "unit_price": r.get("unit_price", ""),
                        "amount": r.get("amount", "")}
                    for r in summary
                ],
                "packages": [
                    {"cml": p["cml"], "name": package_names.get(key, ""),
                     "nw": p["nw"], "gw": p["gw"]}
                    for key, p in packages.items()
                ],
                "total_pkg": len(packages),
                "total_nw": round(total_nw, 3),
                "total_gw": round(total_gw, 3),
                "total_amount": round(invoice_amount, 2),
            })

            grand_pkg += len(packages)
            grand_nw += total_nw
            grand_gw += total_gw
            grand_amount += invoice_amount

        excel_buffer = build_excel(all_results)
        report_id = uuid.uuid4().hex
        report_path = os.path.join(UPLOAD_DIR, f"{report_id}.xlsx")
        with open(report_path, "wb") as out:
            out.write(excel_buffer.getvalue())

        _set_job(
            job_id,
            status="done",
            preview=preview,
            download_id=report_id,
            grand_total_pkg=grand_pkg,
            grand_total_nw=round(grand_nw, 3),
            grand_total_gw=round(grand_gw, 3),
            grand_total_amount=round(grand_amount, 2),
        )
    except Exception as e:
        _set_job(job_id, status="error", error=str(e))


# ---------------------------------------------------------------------------
# PDF PARSING (PKG Counter) - Packing List template
# ---------------------------------------------------------------------------

# CML No.   Volume(m3)   Net Weight(kg)   Gross Weight(kg)
# Strict form: CML No. is immediately followed by the three numbers, with
# nothing else in between, e.g. "STG002608060339 0.005 1.200 1.525"
CML_RE = re.compile(
    r'^(?P<cml>\S+)\s+(?P<vol>[\d.]+)\s+(?P<nw>[\d.,]+)\s+(?P<gw>[\d.,]+)$')

# Looser CML line variant: some templates (e.g. DENSO's "Return Style Code"
# layout) put a free-text field between the CML No. and the trailing
# Volume/Net Weight/Gross Weight numbers, e.g.:
#   "DTG0T11C607240160 PLASTIC PACKAGING (Plastic Slipsheet) 1.156 170.000 170.000"
# This is only tried after CML_RE *and* the item/model line patterns below
# have all failed to match, so it can never swallow an item/model line.
CML_RE_LOOSE = re.compile(
    r'^(?P<cml>\S+)\s+.+?\s+(?P<vol>[\d.]+)\s+(?P<nw>[\d.,]+)\s+(?P<gw>[\d.,]+)$')

# "Normal" template: Customer Order No. Item No. PKG UNIT QTY Total N/W(kg) No. of Cartons
# all on one line, and the order no. looks like "...K123".
ORDER_RE = re.compile(
    r'^(?P<order>\S+K\d+)\s+(?P<itemno>\S+)\s+(?P<unit>\S+)\s+'
    r'(?P<qty>[\d,]+)\s+(?P<totalnw>[\d.,]+)\s+(?P<cartons>\d+)$'
)

# Some templates (e.g. DENSO (Thailand)) put Item No. / UNIT / QTY / Total N/W
# on their own line with NO order number and NO cartons value, e.g.:
#   "TG022108-00509B pcs 60 1.200"
# The Customer Order No. then shows up on the FOLLOWING line together with
# the Description (Customer Item No. is left blank), e.g.:
#   "TG022108-0050 SWITCH THERMO"
ITEM_LINE_RE = re.compile(
    r'^(?P<itemno>\S+)\s+(?P<unit>\S+)\s+(?P<qty>[\d,]+)\s+(?P<totalnw>[\d.,]+)$'
)
ORDER_DESC_RE = re.compile(r'^(?P<order>\S+)\s+(?P<desc>.+)$')

# "Return Style Code" template: Model line has an extra per-unit N/W(kg)
# column before the Total N/W(kg), e.g. "N73T pcs 10 16.700 167.000", with
# only a plain description following (no order/customer-item number), e.g.
# "PLASTIC PACKAGING". Item No. is left blank in this template, so "Model"
# (e.g. "N73T") is used as the item number.
MODEL_LINE_RE = re.compile(
    r'^(?P<itemno>\S+)\s+(?P<unit>\S+)\s+(?P<qty>[\d,]+)\s+'
    r'(?P<nw_unit>[\d.,]+)\s+(?P<totalnw>[\d.,]+)$'
)

# "KN127314-3110 TUBE" -> custitem + desc, both on the same line.
# Require the first token to contain a digit so it looks like an item code
# (otherwise a plain description like "SWITCH THERMO" would get wrongly
# split into custitem="SWITCH" desc="THERMO").
DESC_RE = re.compile(r'^(?P<custitem>\S*\d\S*)\s+(?P<desc>.+)$')

HEADER_PREFIXES = (
    "CML No.", "Customer Order No.", "Customer Item No.",
    "Page", "Total Package", "Model",
    "Phone", "Fax", "Tel", "Tel.", "Email",
)

# Marker used to find the "Ship to" company name on page 1.
SHIP_TO_MARKER_RE = re.compile(r'Ship\s*to', re.IGNORECASE)
COMPANY_HINT_RE = re.compile(r'CO\.,?\s*LTD', re.IGNORECASE)

# "Invoice No." line on the cover page - present on both packing lists and
# invoices in this document family, and is the primary key used to match a
# PKL to its INV. Anchored to line-start so it doesn't match "Ref Invoice
# No." (which has no value on the same line in these templates anyway).
INVOICE_NO_RE = re.compile(r'^Invoice No\.\s*(\S+)')


def extract_ship_to(first_page_text):
    """
    Detects the "Ship to" company / factory name from the raw text of a
    packing list's (or invoice's) first page.

    pdfplumber often merges the "Sold to"/"Consignee" and "Ship to" table
    columns onto the same line (since they sit at the same y-position in
    the PDF), e.g.:

        Ship to Document Information DENSO (THAILAND) CO., LTD.
        DENSO (THAILAND) CO., LTD. Document Type NORMAL ...

    So instead of taking a whole line, we find the "Ship to" marker and
    then grab text up to (and including) the *first* "CO., LTD." style
    company suffix that follows it - that's the factory name, and it
    stops before any of the following merged-in text.

    Returns the company name (e.g. "DENSO (THAILAND) CO., LTD.") or
    "Unknown" if it can't be found, so a PDF with an unrecognised layout
    still gets grouped (under "Unknown") instead of crashing the report.
    """
    text = " ".join(first_page_text.split()
                    )  # collapse all whitespace/newlines

    m = SHIP_TO_MARKER_RE.search(text)
    if not m:
        return "Unknown"

    tail = text[m.end():]
    # "Document Information" is the merged-in column header next to "Ship to" - drop it.
    tail = re.sub(r'^\s*Document Information\s*',
                  '', tail, flags=re.IGNORECASE)

    company_match = re.match(r'\s*(.*?CO\.,?\s*LTD\.?)', tail, re.IGNORECASE)
    if company_match:
        return company_match.group(1).strip()

    return "Unknown"


def extract_invoice_no(first_page_text):
    """Finds the "Invoice No." value on a cover page. Works the same way
    for both packing lists and invoices, since both carry this field."""
    for raw_line in first_page_text.split("\n"):
        m = INVOICE_NO_RE.match(raw_line.strip())
        if m:
            return m.group(1)
    return "Unknown"


def parse_packing_list(pdf_path):
    """
    Parses a packing list PDF exactly as before (CML No. / PKG / weight
    breakdown) - the only addition is also reading the "Invoice No." off
    the cover page, so this shipment can be matched to its invoice later.

    Returns (records, packages, ship_to, invoice_no)

    records    - one dict per order line (used for the item summary sheet).
                 Always has both "custitem" and "desc" filled in as best
                 as possible (desc may be "" if truly not found anywhere).
    packages   - OrderedDict keyed by CML No., each value:
                 {"cml": ..., "vol": float, "nw": float, "gw": float}
                 (used for the combined PKG weight sheet)
    ship_to    - detected Ship To company / factory name (string), used
                 to group the PKG Summary sheet by factory.
    invoice_no - detected Invoice No. (string, "Unknown" if not found),
                 used to match this packing list to its invoice.
    """
    lines = []
    ship_to = "Unknown"
    invoice_no = "Unknown"
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if page_idx == 0:
                ship_to = extract_ship_to(text)
                invoice_no = extract_invoice_no(text)
            for raw_line in text.split("\n"):
                line = raw_line.strip()
                if not line:
                    continue
                if line.startswith(HEADER_PREFIXES):
                    continue
                lines.append(line)
            # pdfplumber caches each page's parsed layout objects (chars,
            # lines, images, etc.) for the lifetime of the `pdf` object.
            # On multi-page PDFs that adds up fast - release it as soon as
            # we're done reading this page's text.
            page.flush_cache()

    records = []
    packages = OrderedDict()
    cur_cml = None
    pending_order = None

    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]

        m = CML_RE.match(line)
        if m:
            cur_cml = m.group("cml")
            packages[cur_cml] = {
                "cml": cur_cml,
                "vol": float(m.group("vol")),
                "nw": float(m.group("nw").replace(",", "")),
                "gw": float(m.group("gw").replace(",", "")),
            }
            i += 1
            continue

        m = ORDER_RE.match(line)
        if m:
            pending_order = m.groupdict()
            i += 1
            continue

        # Template variant: Model line with an extra per-unit N/W(kg) column,
        # e.g. "N73T pcs 10 16.700 167.000", followed by a plain description
        # line with no order/customer-item number, e.g. "PLASTIC PACKAGING".
        m = MODEL_LINE_RE.match(line)
        if m:
            info = m.groupdict()
            info.pop("nw_unit", None)
            info["cartons"] = "0"
            info["order"] = ""
            pending_order = info
            i += 1
            continue

        # Template variant: item info (itemno/unit/qty/totalnw) on its own
        # line, no order no. and no cartons - the order no. + description
        # come together on the next line.
        m = ITEM_LINE_RE.match(line)
        if m:
            item_info = m.groupdict()
            item_info["cartons"] = "0"
            item_info["order"] = ""
            consumed_next = False
            if i + 1 < n:
                nxt = lines[i + 1]
                if not CML_RE.match(nxt) and not CML_RE_LOOSE.match(nxt) and not ORDER_RE.match(nxt) \
                        and not MODEL_LINE_RE.match(nxt) and not ITEM_LINE_RE.match(nxt):
                    om = ORDER_DESC_RE.match(nxt)
                    if om:
                        item_info["order"] = om.group("order")
                        rec = dict(item_info)
                        rec["cml"] = cur_cml
                        rec["custitem"] = ""
                        rec["desc"] = om.group("desc").strip()
                        records.append(rec)
                        i += 2
                        consumed_next = True
            if not consumed_next:
                # No matching order/desc line followed - fall back to the
                # normal pending_order flow so we still capture something
                # instead of dropping the item.
                pending_order = item_info
                i += 1
            continue

        # Loose CML line (CML No. + free-text return style code + vol/nw/gw).
        # Only tried after the strict CML_RE and every item/model/order
        # pattern above have failed, so it can't accidentally swallow an
        # item or model line.
        m = CML_RE_LOOSE.match(line)
        if m:
            cur_cml = m.group("cml")
            packages[cur_cml] = {
                "cml": cur_cml,
                "vol": float(m.group("vol")),
                "nw": float(m.group("nw").replace(",", "")),
                "gw": float(m.group("gw").replace(",", "")),
            }
            i += 1
            continue

        if pending_order is not None:
            custitem = ""
            desc = ""

            m = DESC_RE.match(line)
            if m:
                custitem = m.group("custitem")
                desc = m.group("desc").strip()
                i += 1
            elif " " not in line:
                custitem = line
                i += 1
                if i < n:
                    nxt = lines[i]
                    if not CML_RE.match(nxt) and not CML_RE_LOOSE.match(nxt) and not ORDER_RE.match(nxt) \
                            and not MODEL_LINE_RE.match(nxt) and not ITEM_LINE_RE.match(nxt) \
                            and not nxt.startswith(HEADER_PREFIXES):
                        desc = nxt.strip()
                        i += 1
            else:
                desc = line.strip()
                i += 1

            rec = dict(pending_order)
            rec["cml"] = cur_cml
            rec["custitem"] = custitem
            rec["desc"] = desc
            records.append(rec)
            pending_order = None
            continue

        i += 1

    return records, packages, ship_to, invoice_no


def display_name(custitem, desc):
    desc = (desc or "").strip()
    if desc:
        return desc
    return (custitem or "").strip() or "(unknown)"


def summarize_by_name(records):
    cml_groups = OrderedDict()
    for r in records:
        cml_groups.setdefault(r["cml"], []).append(r)

    summary = OrderedDict()
    package_names = OrderedDict()

    for cml, lines in cml_groups.items():
        within = OrderedDict()
        for r in lines:
            key = (r["itemno"], r["desc"])
            if key not in within:
                within[key] = {
                    "itemno": r["itemno"], "desc": r["desc"], "custitem": r["custitem"],
                    "unit": r["unit"], "qty": 0, "netweight": 0.0,
                }
            within[key]["qty"] += int(r["qty"].replace(",", ""))
            within[key]["netweight"] += float(r["totalnw"].replace(",", ""))

        winner_key = max(within.keys(), key=lambda k: within[k]["netweight"])
        winner = within[winner_key]
        package_names[cml] = display_name(winner["custitem"], winner["desc"])

        for key, data in within.items():
            if key not in summary:
                summary[key] = {
                    "itemno": data["itemno"], "custitem": data["custitem"],
                    "desc": data["desc"], "unit": data["unit"],
                    "count": 0, "total_qty": 0,
                }
            summary[key]["total_qty"] += data["qty"]
            if key == winner_key:
                summary[key]["count"] += 1

    return list(summary.values()), package_names


# ---------------------------------------------------------------------------
# PDF PARSING (PKG Counter) - Invoice (INV) template
# ---------------------------------------------------------------------------
#
# DENSO-style invoices consist of:
#   - a cover page with Invoice No., Ship to, No. of PKG, Net/Gross Weight,
#     Volume, and a Total Amount (Net Amount + Freight + Insurance).
#   - one or more "INVOICE ATTACHED SHEET" pages listing line items, each
#     spread across a fixed 3-line group as pdfplumber extracts it:
#       Row A: "<Customer Order No. or Model>  <Customer Item No. or dims>  <Country>"
#       Row B: "<Description>  <Currency>"
#       Row C: "[Item No.]  <Unit>  <Qty>  <Unit Price>  <Amount>"
#     (Item No. on Row C is only present in some templates - when it's
#     missing, the first token of Row A doubles as the item code, mirroring
#     the same kind of "data lives in a different column depending on
#     template" fallback used in the packing-list parser above.)

# Lines to ignore on "INVOICE ATTACHED SHEET" pages - table headers,
# pagination, and the trailing "Total QTY ... Total Amount ..." line.
INVOICE_LINE_SKIP_PREFIXES = (
    "Page", "INVOICE ATTACHED SHEET", "Invoice No.", "Customer Order No.",
    "No.", "Description", "Model", "Item No.", "Total QTY",
)

# Row C: optional Item No., then Unit (letters only) Qty UnitPrice Amount.
# Anchored start-to-end so it can't accidentally match a header/footer line
# that happens to contain some of these tokens.
ITEM_PRICE_RE = re.compile(
    r'^(?:(?P<itemno>\S+)\s+)?(?P<unit>[A-Za-z]+)\s+(?P<qty>[\d,]+)\s+'
    r'(?P<unitprice>[\d.,]+)\s+(?P<amount>[\d.,]+)$'
)

PKG_COUNT_RE = re.compile(r'No\. of PKG\s+(\d+)')
NET_WEIGHT_RE = re.compile(r'Net Weight\s*\(kg\)\s+([\d,.]+)')
GROSS_WEIGHT_RE = re.compile(r'Gross Weight\s*\(kg\)\s+([\d,.]+)')
VOLUME_RE = re.compile(r'Volume\s*\(m3\)\s+([\d,.]+)')
# The cover page's own Total Amount line, e.g. "Total Amount USD 1,035.90".
# Anchored to the start of the line so it never matches the "Total QTY ...
# Total Amount ..." footer line on the attached-sheet pages.
TOTAL_AMOUNT_RE = re.compile(
    r'^Total Amount\s+(?P<cur>[A-Z]{3})\s+(?P<amt>[\d,.]+)$')


def detect_doc_type(pdf_path):
    """
    Cheap check of a PDF's first couple of pages to decide which parser to
    use. Packing lists have a "CML No." column (per-pallet weight
    breakdown); invoices don't have that, but do have "Unit Price" /
    "Total Amount". Falls back to "packing_list" (the original, default
    behaviour of this app) if neither marker is clearly detected.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            sample_text = ""
            for page in pdf.pages[:2]:
                sample_text += (page.extract_text() or "") + "\n"
                page.flush_cache()
    except Exception:
        return "packing_list"

    if "CML No." in sample_text:
        return "packing_list"
    if "Unit Price" in sample_text and "Total Amount" in sample_text:
        return "invoice"
    return "packing_list"


def parse_invoice(pdf_path):
    """
    Parses a DENSO-style INVOICE PDF.

    Returns (records, header):
      records - one dict per line item:
                {itemno, order, custitem, desc, country, unit, qty,
                 unit_price, amount}
      header  - {invoice_no, ship_to, pkg_count, net_weight, gross_weight,
                 volume, currency, total_amount}
    """
    header = {
        "invoice_no": "Unknown", "ship_to": "Unknown", "pkg_count": 0,
        "net_weight": 0.0, "gross_weight": 0.0, "volume": 0.0,
        "currency": "USD", "total_amount": 0.0,
    }
    item_lines = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""

            if page_idx == 0:
                header["ship_to"] = extract_ship_to(text)
                header["invoice_no"] = extract_invoice_no(text)
                for raw_line in text.split("\n"):
                    line = raw_line.strip()
                    if not line:
                        continue

                    m = PKG_COUNT_RE.search(line)
                    if m:
                        header["pkg_count"] = int(m.group(1))
                        continue
                    m = NET_WEIGHT_RE.search(line)
                    if m:
                        header["net_weight"] = float(
                            m.group(1).replace(",", ""))
                        continue
                    m = GROSS_WEIGHT_RE.search(line)
                    if m:
                        header["gross_weight"] = float(
                            m.group(1).replace(",", ""))
                        continue
                    m = VOLUME_RE.search(line)
                    if m:
                        header["volume"] = float(m.group(1).replace(",", ""))
                        continue
                    m = TOTAL_AMOUNT_RE.match(line)
                    if m:
                        header["currency"] = m.group("cur")
                        header["total_amount"] = float(
                            m.group("amt").replace(",", ""))
                        continue
            else:
                for raw_line in text.split("\n"):
                    line = raw_line.strip()
                    if not line:
                        continue
                    if line.startswith(INVOICE_LINE_SKIP_PREFIXES):
                        continue
                    item_lines.append(line)

            page.flush_cache()

    records = []
    # Item rows come in fixed triples: [order/model + country],
    # [description + currency], [itemno? + unit + qty + unit price + amount]
    for i in range(0, len(item_lines) - 2, 3):
        row_a, row_b, row_c = item_lines[i], item_lines[i + 1], item_lines[i + 2]

        m = ITEM_PRICE_RE.match(row_c)
        if not m:
            # Layout didn't match what we expected for this triple - skip
            # it rather than mis-attributing numbers to the wrong item.
            continue

        a_tokens = row_a.rsplit(" ", 1)
        country = a_tokens[1] if len(a_tokens) == 2 else ""
        a_rest = a_tokens[0] if len(a_tokens) == 2 else row_a
        a_parts = a_rest.split(None, 1)
        order = a_parts[0] if a_parts else ""
        custitem = a_parts[1] if len(a_parts) > 1 else ""

        b_tokens = row_b.rsplit(" ", 1)
        desc = b_tokens[0] if len(b_tokens) == 2 else row_b

        # Some templates leave Item No. blank on Row C - in that case the
        # first token of Row A (e.g. a Model code like "N60T") is the item
        # identifier instead.
        itemno = m.group("itemno") or order

        records.append({
            "itemno": itemno,
            "order": order,
            "custitem": custitem,
            "desc": desc.strip(),
            "country": country,
            "unit": m.group("unit"),
            "qty": int(m.group("qty").replace(",", "")),
            "unit_price": float(m.group("unitprice").replace(",", "")),
            "amount": float(m.group("amount").replace(",", "")),
        })

    return records, header


# ---------------------------------------------------------------------------
# MATCHING - combine a packing list and its invoice into one entry
# ---------------------------------------------------------------------------

# Fallback shipment ID when a PDF's "Invoice No." field can't be found: the
# longest run of 5+ digits in the filename. DENSO's own naming convention
# (e.g. "INV_TG0_651599_20260813094815.pdf") embeds the invoice number
# there too, so this still lets a PKL and INV match up even if one
# template doesn't carry an explicit "Invoice No." line.
FILENAME_ID_RE = re.compile(r'(\d{5,})')


def _shipment_key(parsed_file):
    invoice_no = parsed_file.get("invoice_no")
    if invoice_no and invoice_no != "Unknown":
        return f'inv:{invoice_no}'
    m = FILENAME_ID_RE.search(parsed_file["filename"])
    if m:
        return f'fname:{m.group(1)}'
    return f'file:{id(parsed_file)}'


def summarize_invoice_only(records):
    """
    Builds the same (rows, packages, package_names) shape used for packing
    lists, for an invoice that has no matching PKL - so nothing is lost,
    it just gets its own sheet instead of being merged into one. Each line
    item becomes one row/one pseudo-"package" (vol/nw/gw left at 0, since
    an invoice alone doesn't report per-item weight), carrying its own
    Unit Price / Amount.
    """
    rows = []
    packages = OrderedDict()
    package_names = OrderedDict()

    for idx, r in enumerate(records):
        rows.append({
            "itemno": r["itemno"], "custitem": r["custitem"], "desc": r["desc"],
            "unit": r["unit"], "count": 1, "total_qty": r["qty"],
            "unit_price": r["unit_price"], "amount": r["amount"],
        })
        key = f'{r["itemno"]}#{idx}'
        packages[key] = {"cml": r["itemno"],
                         "vol": 0.0, "nw": 0.0, "gw": 0.0}
        package_names[key] = r["desc"] or r["itemno"]

    return rows, packages, package_names


def build_merged_entries(parsed_files):
    """
    Groups parsed PKL/INV files by shipment (see _shipment_key) and merges
    each matching PKL + INV pair into ONE entry: the packing list's normal
    item/package data, with Unit Price + Amount added to each item row by
    matching Item No. against the invoice's line items (summed if the
    invoice lists the same item more than once, e.g. under two different
    Customer Order Nos.).

    Returns all_results: list of
        (filename, summary_rows, packages, package_names, ship_to, invoice_amount)
    - the same shape build_excel() and the preview builder expect.
    invoice_amount is the shipment's total invoiced value (for the PKG
    Summary subtotal row); it's 0 for a PKL with no matching invoice.
    """
    groups = OrderedDict()
    for pf in parsed_files:
        groups.setdefault(_shipment_key(pf), []).append(pf)

    all_results = []

    for key, items in groups.items():
        pl_items = [p for p in items if p["type"] == "packing_list"]
        inv_items = [p for p in items if p["type"] == "invoice"]

        if pl_items:
            # One sheet per shipment: base it on the packing list (if more
            # than one PKL somehow shares a shipment key, only the first is
            # used - that shouldn't normally happen).
            base = pl_items[0]
            summary = [dict(row) for row in base["summary"]]
            packages = base["packages"]
            package_names = base["package_names"]
            ship_to = base["ship_to"]
            if ship_to == "Unknown" and inv_items:
                ship_to = inv_items[0]["ship_to"]

            invoice_by_itemno = OrderedDict()
            invoice_amount = 0.0
            for inv in inv_items:
                for r in inv["records"]:
                    agg = invoice_by_itemno.setdefault(
                        r["itemno"], {"qty": 0, "amount": 0.0})
                    agg["qty"] += r["qty"]
                    agg["amount"] += r["amount"]
                    invoice_amount += r["amount"]

            for row in summary:
                match = invoice_by_itemno.get(row["itemno"])
                if match:
                    row["unit_price"] = round(
                        match["amount"] / match["qty"], 4) if match["qty"] else 0
                    row["amount"] = round(match["amount"], 2)
                else:
                    row["unit_price"] = ""
                    row["amount"] = ""

            all_results.append(
                (base["filename"], summary, packages, package_names,
                 ship_to, invoice_amount))
        else:
            # No matching packing list - keep the invoice(s) as their own
            # sheet(s) so the data still shows up somewhere.
            for inv in inv_items:
                summary, packages, package_names = summarize_invoice_only(
                    inv["records"])
                invoice_amount = sum(r["amount"] for r in inv["records"])
                all_results.append(
                    (inv["filename"], summary, packages, package_names,
                     inv["ship_to"], invoice_amount))

    return all_results


# ---------------------------------------------------------------------------
# EXCEL EXPORT (PKG Counter)
# ---------------------------------------------------------------------------

def safe_sheet_name(name, used_names):
    base = re.sub(r'[\\/*?:\[\]]', "_", name)[:31]
    candidate = base
    i = 1
    while candidate in used_names:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used_names.add(candidate)
    return candidate


def style_header(ws, headers, header_fill, header_font):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, headers, rows_as_strs):
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in rows_as_strs:
            max_len = max(max_len, len(row[col_idx - 1]))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4


def build_excel(all_results):
    """
    all_results: list of tuples
        (pdf_name, item_rows, packages, package_names, ship_to, invoice_amount)

    item_rows may carry "unit_price"/"amount" (blank string if no matching
    invoice was found for that item) - that's the only content change to
    the per-shipment item sheet vs. before: two extra columns, Unit Price
    and Amount. invoice_amount (the shipment's total invoiced value) feeds
    the Amount (USD) column on the PKG Summary subtotal row.
    """
    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    header_fill = PatternFill(start_color="305496",
                              end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill(start_color="D9E1F2",
                           end_color="D9E1F2", fill_type="solid")
    sub_font = Font(bold=True)
    factory_fill = PatternFill(start_color="8EA9DB",
                               end_color="8EA9DB", fill_type="solid")
    factory_font = Font(bold=True)
    section_fill = PatternFill(start_color="203864",
                               end_color="203864", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF")
    grand_fill = PatternFill(start_color="305496",
                             end_color="305496", fill_type="solid")
    grand_font = Font(bold=True, color="FFFFFF")

    item_headers = ["Item No.", "Customer Item No.", "Description", "Unit",
                    "PKG (Count)", "Total Qty", "Unit Price", "Amount"]

    # ---- Per-shipment item summary sheets ----
    for pdf_name, rows, packages, package_names, ship_to, invoice_amount in all_results:
        base_title = os.path.splitext(pdf_name)[0]
        sheet_title = safe_sheet_name(base_title, used_names)
        ws = wb.create_sheet(title=sheet_title)
        style_header(ws, item_headers, header_fill, header_font)
        for row in rows:
            desc_display = display_name(row["custitem"], row["desc"])
            ws.append([
                row["itemno"], row["custitem"], desc_display,
                row["unit"], row["count"], row["total_qty"],
                row.get("unit_price", ""), row.get("amount", ""),
            ])
        row_strs = [
            [row["itemno"], row["custitem"], display_name(row["custitem"], row["desc"]),
             row["unit"], str(row["count"]), str(row["total_qty"]),
             (f'{row["unit_price"]:.4f}' if isinstance(row.get("unit_price"), (int, float)) else ""),
             (f'{row["amount"]:.2f}' if isinstance(row.get("amount"), (int, float)) else "")]
            for row in rows
        ]
        autosize(ws, item_headers, row_strs)
        ws.freeze_panes = "A2"

    # ---- PKG Summary sheet, grouped by Ship To (factory), then shipment ----
    # Column G (Amount (USD)) is the only new column vs. the original app -
    # it carries each shipment's total invoiced amount on the Subtotal
    # row (blank on individual CML/package rows, since price isn't known
    # at that granularity).
    pkg_headers = ["Ship To", "Invoice", "CML No.", "Description",
                   "Net Weight (kg)", "Gross Weight (kg)", "Amount (USD)"]
    ws_pkg = wb.create_sheet(title="PKG Summary")
    style_header(ws_pkg, pkg_headers, header_fill, header_font)

    # Group results by detected Ship To, preserving first-seen order.
    grouped_by_ship_to = OrderedDict()
    for pdf_name, rows, packages, package_names, ship_to, invoice_amount in all_results:
        grouped_by_ship_to.setdefault(ship_to, []).append(
            (pdf_name, packages, package_names, invoice_amount))

    pkg_row_strs = []
    grand_nw = 0.0
    grand_gw = 0.0
    grand_amount = 0.0
    grand_pkg_count = 0

    for ship_to, items in grouped_by_ship_to.items():
        # --- Factory section header row ---
        section_row_idx = ws_pkg.max_row + 1
        ws_pkg.append([f"Ship To: {ship_to}", "", "", "", "", "", ""])
        ws_pkg.merge_cells(start_row=section_row_idx, start_column=1,
                           end_row=section_row_idx, end_column=len(pkg_headers))
        for col_idx in range(1, len(pkg_headers) + 1):
            cell = ws_pkg.cell(row=section_row_idx, column=col_idx)
            cell.fill = section_fill
            cell.font = section_font
        pkg_row_strs.append([f"Ship To: {ship_to}", "", "", "", "", "", ""])

        ship_nw = 0.0
        ship_gw = 0.0
        ship_amount = 0.0
        ship_pkg_count = 0

        for pdf_name, packages, package_names, invoice_amount in items:
            invoice_label = os.path.splitext(pdf_name)[0]
            inv_nw = 0.0
            inv_gw = 0.0

            for key, pkg in packages.items():
                name = package_names.get(key, "")
                ws_pkg.append([ship_to, invoice_label, pkg["cml"], name,
                               pkg["nw"], pkg["gw"], ""])
                pkg_row_strs.append(
                    [ship_to, invoice_label, pkg["cml"], name,
                     f'{pkg["nw"]:.3f}', f'{pkg["gw"]:.3f}', ""])
                inv_nw += pkg["nw"]
                inv_gw += pkg["gw"]

            sub_row_idx = ws_pkg.max_row + 1
            ws_pkg.append([ship_to, invoice_label,
                           f"Subtotal ({len(packages)} pkg)", "",
                           round(inv_nw, 3), round(inv_gw, 3),
                           round(invoice_amount, 2) if invoice_amount else ""])
            for col_idx in range(1, len(pkg_headers) + 1):
                cell = ws_pkg.cell(row=sub_row_idx, column=col_idx)
                cell.fill = sub_fill
                cell.font = sub_font
            pkg_row_strs.append(
                [ship_to, invoice_label, f"Subtotal ({len(packages)} pkg)", "",
                 f"{inv_nw:.3f}", f"{inv_gw:.3f}",
                 f"{invoice_amount:.2f}" if invoice_amount else ""])

            ship_nw += inv_nw
            ship_gw += inv_gw
            ship_amount += invoice_amount
            ship_pkg_count += len(packages)

        # --- Factory (Ship To) subtotal row ---
        factory_row_idx = ws_pkg.max_row + 1
        ws_pkg.append(["", f"{ship_to} Total ({ship_pkg_count} pkg)", "", "",
                       round(ship_nw, 3), round(ship_gw, 3),
                       round(ship_amount, 2) if ship_amount else ""])
        for col_idx in range(1, len(pkg_headers) + 1):
            cell = ws_pkg.cell(row=factory_row_idx, column=col_idx)
            cell.fill = factory_fill
            cell.font = factory_font
        pkg_row_strs.append(
            ["", f"{ship_to} Total ({ship_pkg_count} pkg)", "", "",
             f"{ship_nw:.3f}", f"{ship_gw:.3f}",
             f"{ship_amount:.2f}" if ship_amount else ""])

        grand_nw += ship_nw
        grand_gw += ship_gw
        grand_amount += ship_amount
        grand_pkg_count += ship_pkg_count

    # --- Grand total row across all factories ---
    grand_row_idx = ws_pkg.max_row + 1
    ws_pkg.append(["", f"Grand Total ({grand_pkg_count} pkg)", "", "",
                   round(grand_nw, 3), round(grand_gw, 3),
                   round(grand_amount, 2) if grand_amount else ""])
    for col_idx in range(1, len(pkg_headers) + 1):
        cell = ws_pkg.cell(row=grand_row_idx, column=col_idx)
        cell.fill = grand_fill
        cell.font = grand_font
    pkg_row_strs.append(
        ["", f"Grand Total ({grand_pkg_count} pkg)", "", "",
         f"{grand_nw:.3f}", f"{grand_gw:.3f}",
         f"{grand_amount:.2f}" if grand_amount else ""])

    autosize(ws_pkg, pkg_headers, pkg_row_strs)
    ws_pkg.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# ROUTES - PKG Counter
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/count", methods=["POST"])
def count_pkg():
    files = request.files.getlist("pdfs")
    if not files or files[0].filename == "":
        return jsonify({"error": "No files uploaded."}), 400

    # Save every uploaded PDF to disk right away - this is fast (no parsing
    # yet), so the HTTP request can return almost instantly instead of
    # staying open for however long parsing the whole batch takes. PKL and
    # INV files can be mixed freely in the same upload - matching PDFs for
    # the same shipment are merged into one sheet in the background job.
    saved_files = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        temp_path = os.path.join(
            UPLOAD_DIR, f"{uuid.uuid4().hex}_{f.filename}")
        f.save(temp_path)
        saved_files.append((temp_path, f.filename))

    if not saved_files:
        return jsonify({"error": "No valid PDF files uploaded."}), 400

    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "queued",
            "progress_done": 0,
            "progress_total": len(saved_files),
            "current_file": None,
        }

    thread = threading.Thread(
        target=_process_count_job, args=(job_id, saved_files), daemon=True)
    thread.start()

    return jsonify({"job_id": job_id, "total_files": len(saved_files)})


@app.route("/count/status/<job_id>")
def count_status(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if job is None:
        return jsonify({"error": "Unknown job ID."}), 404
    # Return a plain copy so we're not holding the lock while jsonify runs.
    return jsonify(job)


@app.route("/download/<report_id>")
def download(report_id):
    report_path = os.path.join(UPLOAD_DIR, f"{report_id}.xlsx")
    if not os.path.isfile(report_path):
        return "Report not found or expired.", 404
    return send_file(
        report_path,
        as_attachment=True,
        download_name="PKG_Count_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# ROUTES - Combine PDFs
# ---------------------------------------------------------------------------

@app.route("/combine")
def combine_page():
    return render_template("combine.html")


@app.route("/combine", methods=["POST"])
def combine_pdfs_route():
    files = request.files.getlist("pdfs")
    if not files or files[0].filename == "":
        return jsonify({"error": "No files uploaded."}), 400

    batch_id = uuid.uuid4().hex
    batch_folder = os.path.join(UPLOAD_DIR, f"combine_{batch_id}")
    os.makedirs(batch_folder, exist_ok=True)

    for f in files:
        if f.filename.lower().endswith(".pdf"):
            f.save(os.path.join(batch_folder, f.filename))

    output_path = os.path.join(UPLOAD_DIR, f"combined_{batch_id}.pdf")

    try:
        result = combine(batch_folder, output_path)
    except ValueError as e:
        shutil.rmtree(batch_folder, ignore_errors=True)
        return jsonify({"error": str(e)}), 400

    shutil.rmtree(batch_folder, ignore_errors=True)

    return jsonify({
        "invoice_count": result["invoice_count"],
        "details": result["details"],
        "unclassified": result["unclassified"],
        "download_id": f"combined_{batch_id}",
    })


@app.route("/download_combine/<download_id>")
def download_combine(download_id):
    file_path = os.path.join(UPLOAD_DIR, f"{download_id}.pdf")
    if not os.path.isfile(file_path):
        return "File not found or expired.", 404
    return send_file(
        file_path,
        as_attachment=True,
        download_name="combined_output.pdf",
        mimetype="application/pdf",
    )


if __name__ == "__main__":
    # Local dev only - Render runs this via gunicorn instead (see Start
    # Command in the docstring at the top of this file), so debug=True here
    # never reaches production.
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)), debug=True)